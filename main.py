from openpyxl import load_workbook

wb2 = load_workbook('abc.xlsx')
#wb2 = load_workbook('Hoja simplificada calculo anual.xlsx')
print(wb2.sheetnames)

ws = wb2['Hoja1']
c1 = ws['A1']
c2 = ws['A2']
c3 = ws['A3']

print(c1.value)
print(c2.value)
print(c3.value)

#print(dir(c1))